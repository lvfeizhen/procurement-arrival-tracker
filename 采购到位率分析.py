"""
采购到位率查询系统
功能：
1. 从 PurchaseOrder 表中筛选出与 PR 表"单据编号"匹配的数据
2. 将筛选结果与采购承诺导出数据按 (单据编号+行号) ↔ (U9采购单号+U9采购单行号) 匹配
3. 计算到位状态：未到位 / 部分未到位 / 已到位
4. 输出统一格式的 Excel 报表
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import glob
import warnings
warnings.filterwarnings('ignore')

# ============================================================
# 1. 自动识别并读取文件
# ============================================================
BASE_DIR = r"C:\Users\Administrator\Desktop\采购到位率"

def find_file(pattern_keyword, exclude_keyword=None):
    """根据关键字查找文件"""
    files = [f for f in os.listdir(BASE_DIR)
             if f.endswith(('.xlsx', '.xls'))
             and not f.startswith('~$')      # 排除Excel临时文件
             and pattern_keyword in f]
    if exclude_keyword:
        files = [f for f in files if exclude_keyword not in f]
    if not files:
        raise FileNotFoundError(f"未找到包含 '{pattern_keyword}' 的文件")
    return os.path.join(BASE_DIR, files[0])

# 识别 PR 报表
pr_file = find_file('PR')
print(f"[1/6] 读取 PR 报表: {os.path.basename(pr_file)}")
df_pr = pd.read_excel(pr_file, header=1)  # 第2行为表头
print(f"      行数: {len(df_pr)}, 列数: {len(df_pr.columns)}")

# 识别 PurchaseOrder 报表
po_file = find_file('PurchaseOrder')
print(f"[2/6] 读取 PurchaseOrder 报表: {os.path.basename(po_file)}")
df_po = pd.read_excel(po_file, header=1)  # 第2行为表头
print(f"      行数: {len(df_po)}, 列数: {len(df_po.columns)}")

# 识别 采购承诺导出数据
commit_file = find_file('采购承诺导出数据')
print(f"[3/6] 读取 采购承诺导出数据: {os.path.basename(commit_file)}")
df_commit = pd.read_excel(commit_file, header=0)  # 第1行为表头
print(f"      行数: {len(df_commit)}, 列数: {len(df_commit.columns)}")

# ============================================================
# 2. 数据清洗与标准化
# ============================================================

# --- PR 表：提取"单据编号"集合 ---
pr_docs = set(df_pr['单据编号'].dropna().astype(str).str.strip())
print(f"\n[4/6] PR 表中唯一单据编号数: {len(pr_docs)}")

# --- PurchaseOrder 表 ---
# 重命名关键列以便处理
df_po = df_po.rename(columns={
    '来源单据号': '来源系统单据号',
    '单据编号': 'PO单据编号',
    '行号': 'PO行号',
    '采购数量': '采购数量',
    '累计实收数量': '累计实收数量',
})

# 确保关键列类型正确
df_po['来源系统单据号'] = df_po['来源系统单据号'].astype(str).str.strip()
df_po['PO单据编号'] = df_po['PO单据编号'].astype(str).str.strip()
df_po['PO行号'] = df_po['PO行号'].apply(lambda x: str(int(x)) if pd.notna(x) else '')
df_po['采购数量'] = pd.to_numeric(df_po['采购数量'], errors='coerce').fillna(0)
df_po['累计实收数量'] = pd.to_numeric(df_po['累计实收数量'], errors='coerce').fillna(0)

# --- 采购承诺导出数据 ---
df_commit = df_commit.rename(columns={
    'U9采购单号': 'U9采购单号',
    'U9采购单行号': 'U9采购单行号',
    '承诺收货日期': '承诺收货日期',
    '承诺收货数量': '承诺收货数量',
})

df_commit['U9采购单号'] = df_commit['U9采购单号'].astype(str).str.strip()
df_commit['U9采购单行号'] = df_commit['U9采购单行号'].apply(lambda x: str(int(x)) if pd.notna(x) else '')
df_commit['承诺收货数量'] = pd.to_numeric(df_commit['承诺收货数量'], errors='coerce').fillna(0)

# 处理承诺收货日期
df_commit['承诺收货日期'] = pd.to_datetime(df_commit['承诺收货日期'], errors='coerce')

# ============================================================
# 3. 第一步筛选：PurchaseOrder 中"来源系统单据号"在 PR 表"单据编号"中
# ============================================================
print(f"\n      筛选前 PurchaseOrder 行数: {len(df_po)}")
df_po_filtered = df_po[df_po['来源系统单据号'].isin(pr_docs)].copy()
print(f"      筛选后 PurchaseOrder 行数: {len(df_po_filtered)}")

# ============================================================
# 4. 对 PurchaseOrder 按 (PO单据编号 + PO行号) 分组求和累计实收数量
# ============================================================
# 同一采购单行可能有多条收货记录，需要聚合
df_po_agg = df_po_filtered.groupby(
    ['PO单据编号', 'PO行号'], as_index=False
).agg({
    '采购数量': 'first',           # 采购数量应该一致，取第一个
    '累计实收数量': 'sum',          # 累计实收数量求和
    '来源系统单据号': 'first',      # 保留来源PR号
    '供应商': 'first',             # 保留供应商信息
    '料号': 'first',              # 保留料号
    '料品': 'first',              # 保留料品名称
    '料品规格': 'first',           # 保留料品规格
    '计价单位': 'first',           # 保留单位
    '单据状态': 'first',           # 保留状态
    '创建时间': 'first',           # 保留创建时间
})

print(f"      聚合后 PurchaseOrder 行数（按单据编号+行号去重）: {len(df_po_agg)}")

# ============================================================
# 5. 第二步匹配：采购承诺 (U9采购单号 + U9采购单行号) ↔ PurchaseOrder (PO单据编号 + PO行号)
# ============================================================
print(f"\n[5/6] 执行数据匹配...")

# 创建匹配键
df_po_agg['_match_key'] = df_po_agg['PO单据编号'] + '|' + df_po_agg['PO行号']
df_commit['_match_key'] = df_commit['U9采购单号'] + '|' + df_commit['U9采购单行号']

# 左连接：以采购承诺数据为主表
df_result = df_commit.merge(
    df_po_agg,
    on='_match_key',
    how='left',
    suffixes=('_承诺', '_PO')
)

print(f"      采购承诺数据总行数: {len(df_commit)}")
print(f"      匹配成功行数: {df_result['PO单据编号'].notna().sum()}")
print(f"      匹配失败行数: {df_result['PO单据编号'].isna().sum()}")

# 分离匹配失败的数据（来源PR单号为空），单独保存
df_unmatched = df_result[df_result['PO单据编号'].isna()].copy()
# 明细表只保留匹配成功的数据
df_result = df_result[df_result['PO单据编号'].notna()].copy()
print(f"      剔除匹配失败后保留行数: {len(df_result)}")

# ============================================================
# 6. 计算到位状态
# ============================================================
print(f"\n[6/6] 计算到位状态...")

def calc_arrival_status(row):
    """
    判断到位状态：
    - 采购数量 - 累计实收数量 == 采购数量 (即累计实收==0) → 未到位
    - 采购数量 - 累计实收数量 > 0 且累计实收 > 0 → 部分未到位
    - 采购数量 - 累计实收数量 <= 0 → 已到位
    """
    qty = row['采购数量']
    received = row['累计实收数量']

    if pd.isna(qty) or pd.isna(received) or qty == 0:
        return '未到位'

    unreceived = qty - received

    if received == 0:
        return '未到位'
    elif unreceived > 0:
        return '部分未到位'
    else:  # unreceived <= 0
        return '已到位'

df_result['采购数量'] = df_result['采购数量'].fillna(0)
df_result['累计实收数量'] = df_result['累计实收数量'].fillna(0)
df_result['未到位数量'] = df_result['采购数量'] - df_result['累计实收数量']
df_result['到位状态'] = df_result.apply(calc_arrival_status, axis=1)

# 到位率计算（已到位=100%，其他按比例）
def calc_arrival_rate(row):
    qty = row['采购数量']
    received = row['累计实收数量']
    if pd.isna(qty) or qty == 0:
        return 0
    rate = received / qty
    return round(min(rate, 1.0), 4)  # 最大100%

df_result['到位率'] = df_result.apply(calc_arrival_rate, axis=1)

# ============================================================
# 7. 整理输出列
# ============================================================
# 选择并重命名输出列
output_columns = {
    'U9采购单号': 'U9采购单号',
    'U9采购单行号': 'U9采购单行号',
    '来源系统单据号': '来源PR单号',
    'PO单据编号': '采购订单号',
    'PO行号': '采购订单行号',
    '承诺收货日期': '承诺收货日期',
    '承诺收货数量': '承诺收货数量',
    '采购数量': '采购数量',
    '累计实收数量': '累计实收数量',
    '未到位数量': '未到位数量',
    '到位率': '到位率',
    '到位状态': '到位状态',
    '供应商': '供应商',
    '料号': '料号',
    '料品': '料品',
    '料品规格': '料品规格',
    '计价单位': '计价单位',
    '单据状态': '采购订单状态',
    '业务员姓名': '业务员',
    '物料规格': '物料规格',
    '物料型号': '物料型号',
    '物料代码': '物料代码',
}

# 只保留存在的列
existing_cols = {k: v for k, v in output_columns.items() if k in df_result.columns}
df_output = df_result[list(existing_cols.keys())].rename(columns=existing_cols)

# 按到位状态排序：未到位 > 部分未到位 > 已到位
status_order = {'未到位': 0, '部分未到位': 1, '已到位': 2}
df_output['_sort'] = df_output['到位状态'].map(status_order)
df_output = df_output.sort_values(['_sort', '承诺收货日期']).drop(columns=['_sort'])

# ============================================================
# 8. 输出 Excel 文件
# ============================================================
output_file = os.path.join(BASE_DIR, f'采购到位率分析结果_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # --- Sheet 1: 到位率明细 ---
    df_output.to_excel(writer, sheet_name='到位率明细', index=False)

    # --- Sheet 2: 汇总统计 ---
    summary_data = []

    # 按到位状态统计
    for status in ['未到位', '部分未到位', '已到位']:
        subset = df_output[df_output['到位状态'] == status]
        summary_data.append({
            '到位状态': status,
            '行数': len(subset),
            '采购数量合计': subset['采购数量'].sum(),
            '累计实收数量合计': subset['累计实收数量'].sum(),
            '未到位数量合计': subset['未到位数量'].sum(),
            '综合到位率': round(
                subset['累计实收数量'].sum() / subset['采购数量'].sum()
                if subset['采购数量'].sum() > 0 else 0, 4
            ),
        })

    # 总计行
    summary_data.append({
        '到位状态': '合计',
        '行数': len(df_output),
        '采购数量合计': df_output['采购数量'].sum(),
        '累计实收数量合计': df_output['累计实收数量'].sum(),
        '未到位数量合计': df_output['未到位数量'].sum(),
        '综合到位率': round(
            df_output['累计实收数量'].sum() / df_output['采购数量'].sum()
            if df_output['采购数量'].sum() > 0 else 0, 4
        ),
    })

    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='汇总统计', index=False)

    # --- Sheet 3: 匹配失败的数据（采购承诺中有但 PO 中找不到的） ---
    if len(df_unmatched) > 0:
        unmatched_cols = {
            'U9采购单号': 'U9采购单号',
            'U9采购单行号': 'U9采购单行号',
            '承诺收货日期': '承诺收货日期',
            '承诺收货数量': '承诺收货数量',
            '业务员姓名': '业务员',
            '物料规格': '物料规格',
            '物料型号': '物料型号',
            '物料代码': '物料代码',
        }
        existing_unmatched = {k: v for k, v in unmatched_cols.items() if k in df_unmatched.columns}
        df_unmatched_out = df_unmatched[list(existing_unmatched.keys())].rename(columns=existing_unmatched)
        df_unmatched_out.to_excel(writer, sheet_name='匹配失败数据', index=False)

    # --- Sheet 4: 到期履约情况（承诺收货日期 <= 今天，含全部到位状态） ---
    today = pd.Timestamp(datetime.now().date())
    df_overdue = df_output[
        (df_output['承诺收货日期'].notna()) &
        (df_output['承诺收货日期'] <= today)
    ].copy()

    if len(df_overdue) > 0:
        # 按到位状态排序：未到位 > 部分未到位 > 已到位，同状态按承诺日期升序
        overdue_sort = {'未到位': 0, '部分未到位': 1, '已到位': 2}
        df_overdue['_sort'] = df_overdue['到位状态'].map(overdue_sort)
        df_overdue = df_overdue.sort_values(['_sort', '承诺收货日期']).drop(columns=['_sort'])
        df_overdue.to_excel(writer, sheet_name='到期履约情况', index=False)

        # 到期汇总
        overdue_summary = []
        for status in ['未到位', '部分未到位', '已到位']:
            subset = df_overdue[df_overdue['到位状态'] == status]
            overdue_summary.append({
                '到位状态': status,
                '行数': len(subset),
                '采购数量合计': subset['采购数量'].sum(),
                '累计实收数量合计': subset['累计实收数量'].sum(),
                '未到位数量合计': subset['未到位数量'].sum(),
                '综合到位率': round(
                    subset['累计实收数量'].sum() / subset['采购数量'].sum()
                    if subset['采购数量'].sum() > 0 else 0, 4
                ),
            })
        overdue_summary.append({
            '到位状态': '合计',
            '行数': len(df_overdue),
            '采购数量合计': df_overdue['采购数量'].sum(),
            '累计实收数量合计': df_overdue['累计实收数量'].sum(),
            '未到位数量合计': df_overdue['未到位数量'].sum(),
            '综合到位率': round(
                df_overdue['累计实收数量'].sum() / df_overdue['采购数量'].sum()
                if df_overdue['采购数量'].sum() > 0 else 0, 4
            ),
        })
        df_overdue_summary = pd.DataFrame(overdue_summary)
        df_overdue_summary.to_excel(writer, sheet_name='到期履约汇总', index=False)
        print(f"\n  到期履约数据: {len(df_overdue)} 条 (承诺收货日期 <= {today.strftime('%Y-%m-%d')})")
    else:
        print(f"\n  无到期数据")

# ============================================================
# 9. 美化格式
# ============================================================
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = load_workbook(output_file)

# 定义样式
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_font = Font(name='微软雅黑', size=10)
data_alignment = Alignment(horizontal='center', vertical='center')

# 到位状态颜色
status_fills = {
    '未到位': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),    # 红色
    '部分未到位': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),  # 黄色
    '已到位': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),      # 绿色
}

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

def format_sheet(ws, status_col_name=None):
    """格式化工作表"""
    # 设置列宽
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header_text = ws.cell(row=1, column=col_idx).value
        if header_text and len(str(header_text)) > 6:
            ws.column_dimensions[col_letter].width = 16
        else:
            ws.column_dimensions[col_letter].width = 14

    # 格式化表头
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 格式化数据行
    status_col_idx = None
    if status_col_name:
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == status_col_name:
                status_col_idx = col_idx
                break

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

        # 到位状态列着色
        if status_col_idx:
            status_cell = ws.cell(row=row_idx, column=status_col_idx)
            status_val = str(status_cell.value) if status_cell.value else ''
            if status_val in status_fills:
                status_cell.fill = status_fills[status_val]

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 自动筛选
    ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

# 格式化"到位率明细"sheet
ws_detail = wb['到位率明细']
format_sheet(ws_detail, status_col_name='到位状态')

# 格式化"汇总统计"sheet
ws_summary = wb['汇总统计']
format_sheet(ws_summary, status_col_name='到位状态')

# 格式化"匹配失败数据"sheet（如果存在）
if '匹配失败数据' in wb.sheetnames:
    ws_unmatched = wb['匹配失败数据']
    format_sheet(ws_unmatched)

# 格式化"到期履约情况"sheet（如果存在）
if '到期履约情况' in wb.sheetnames:
    ws_overdue = wb['到期履约情况']
    format_sheet(ws_overdue, status_col_name='到位状态')

# 格式化"到期履约汇总"sheet（如果存在）
if '到期履约汇总' in wb.sheetnames:
    ws_overdue_summary = wb['到期履约汇总']
    format_sheet(ws_overdue_summary, status_col_name='到位状态')

# 设置百分比格式（对所有含"到位率"列的 sheet）
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rate_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_idx).value == '到位率':
            rate_col_idx = col_idx
            break
    if rate_col_idx:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=rate_col_idx)
            cell.number_format = '0.00%'

wb.save(output_file)

print(f"\n{'='*60}")
print(f"  分析完成！")
print(f"  输出文件: {output_file}")
print(f"{'='*60}")
print(f"\n  --- 汇总统计 ---")
for row in summary_data:
    status = row['到位状态']
    count = row['行数']
    rate = row['综合到位率']
    print(f"  {status}: {count} 条, 综合到位率: {rate:.2%}")
