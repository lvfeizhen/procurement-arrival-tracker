"""
采购到位率查询系统 - Web版
使用方法: 双击 启动.bat 或在命令行运行 python app.py
          或打包后双击 采购到位率查询系统.exe
"""

import os
import sys
import glob
import pandas as pd
import numpy as np
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

from flask import Flask, render_template, request, jsonify, send_file

# PyInstaller 打包兼容处理
if getattr(sys, 'frozen', False):
    # 打包后的 exe 运行环境
    BASE_DIR = os.path.dirname(sys.executable)
    TEMPLATE_DIR = os.path.join(sys._MEIPASS, 'templates')
    app = Flask(__name__, template_folder=TEMPLATE_DIR)
else:
    # 普通 Python 运行环境
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    app = Flask(__name__)

app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB 上限
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(BASE_DIR, 'outputs')

# 确保目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# ============================================================
# 核心分析逻辑
# ============================================================
def analyze(pr_path, po_path, commit_path):
    """执行采购到位率分析，返回 (output_path, summary, overdue_summary)"""

    # 1. 读取文件
    df_pr = pd.read_excel(pr_path, header=1)
    df_po = pd.read_excel(po_path, header=1)
    df_commit = pd.read_excel(commit_path, header=0)

    # 2. PR 单据编号集合
    pr_docs = set(df_pr['单据编号'].dropna().astype(str).str.strip())

    # 3. PurchaseOrder 数据清洗
    df_po = df_po.rename(columns={
        '来源单据号': '来源系统单据号',
        '单据编号': 'PO单据编号',
        '行号': 'PO行号',
        '采购数量': '采购数量',
        '累计实收数量': '累计实收数量',
    })
    df_po['来源系统单据号'] = df_po['来源系统单据号'].astype(str).str.strip()
    df_po['PO单据编号'] = df_po['PO单据编号'].astype(str).str.strip()
    df_po['PO行号'] = df_po['PO行号'].apply(lambda x: str(int(x)) if pd.notna(x) else '')
    df_po['采购数量'] = pd.to_numeric(df_po['采购数量'], errors='coerce').fillna(0)
    df_po['累计实收数量'] = pd.to_numeric(df_po['累计实收数量'], errors='coerce').fillna(0)

    # 4. 采购承诺数据清洗
    df_commit = df_commit.rename(columns={
        'U9采购单号': 'U9采购单号',
        'U9采购单行号': 'U9采购单行号',
        '承诺收货日期': '承诺收货日期',
        '承诺收货数量': '承诺收货数量',
    })
    df_commit['U9采购单号'] = df_commit['U9采购单号'].astype(str).str.strip()
    df_commit['U9采购单行号'] = df_commit['U9采购单行号'].apply(lambda x: str(int(x)) if pd.notna(x) else '')
    df_commit['承诺收货数量'] = pd.to_numeric(df_commit['承诺收货数量'], errors='coerce').fillna(0)
    df_commit['承诺收货日期'] = pd.to_datetime(df_commit['承诺收货日期'], errors='coerce')

    # 5. 筛选：PO 来源单据号 在 PR 单据编号中
    df_po_filtered = df_po[df_po['来源系统单据号'].isin(pr_docs)].copy()

    # 6. 聚合：按 PO单据编号+PO行号 分组
    agg_cols = {
        '采购数量': 'first',
        '累计实收数量': 'sum',
        '来源系统单据号': 'first',
        '供应商': 'first',
        '料号': 'first',
        '料品': 'first',
        '料品规格': 'first',
        '计价单位': 'first',
        '单据状态': 'first',
        '创建时间': 'first',
    }
    available_agg = {k: v for k, v in agg_cols.items() if k in df_po_filtered.columns}
    df_po_agg = df_po_filtered.groupby(['PO单据编号', 'PO行号'], as_index=False).agg(available_agg)

    # 7. 匹配
    df_po_agg['_match_key'] = df_po_agg['PO单据编号'] + '|' + df_po_agg['PO行号']
    df_commit['_match_key'] = df_commit['U9采购单号'] + '|' + df_commit['U9采购单行号']

    df_result = df_commit.merge(df_po_agg, on='_match_key', how='left', suffixes=('_承诺', '_PO'))

    matched_count = df_result['PO单据编号'].notna().sum()
    unmatched_count = df_result['PO单据编号'].isna().sum()

    # 分离匹配失败的数据
    df_unmatched = df_result[df_result['PO单据编号'].isna()].copy()
    df_result = df_result[df_result['PO单据编号'].notna()].copy()

    # 8. 计算到位状态
    def calc_arrival_status(row):
        qty = row['采购数量']
        received = row['累计实收数量']
        if pd.isna(qty) or pd.isna(received) or qty == 0:
            return '未到位'
        unreceived = qty - received
        if received == 0:
            return '未到位'
        elif unreceived > 0:
            return '部分未到位'
        else:
            return '已到位'

    df_result['采购数量'] = df_result['采购数量'].fillna(0)
    df_result['累计实收数量'] = df_result['累计实收数量'].fillna(0)
    df_result['未到位数量'] = df_result['采购数量'] - df_result['累计实收数量']
    df_result['到位状态'] = df_result.apply(calc_arrival_status, axis=1)

    def calc_arrival_rate(row):
        qty = row['采购数量']
        received = row['累计实收数量']
        if pd.isna(qty) or qty == 0:
            return 0
        return round(min(received / qty, 1.0), 4)

    df_result['到位率'] = df_result.apply(calc_arrival_rate, axis=1)

    # 9. 整理输出列
    output_columns = {
        'U9采购单号': 'U9采购单号',
        'U9采购单行号': 'U9采购单行号',
        '来源系统单据号': '来源PR单号',
        'PO单据编号': '采购订单号',
        'PO行号': '采购订单行号',
        '承诺收货日期': '承诺收货日期',
        '承诺收货数量': '承诺收货数量',
        '采购数量': '采购数量',
        '累计实收数量': '累计实收数量',
        '未到位数量': '未到位数量',
        '到位率': '到位率',
        '到位状态': '到位状态',
        '供应商': '供应商',
        '料号': '料号',
        '料品': '料品',
        '料品规格': '料品规格',
        '计价单位': '计价单位',
        '单据状态': '采购订单状态',
        '业务员姓名': '业务员',
        '物料规格': '物料规格',
        '物料型号': '物料型号',
        '物料代码': '物料代码',
    }
    existing_cols = {k: v for k, v in output_columns.items() if k in df_result.columns}
    df_output = df_result[list(existing_cols.keys())].rename(columns=existing_cols)

    status_order = {'未到位': 0, '部分未到位': 1, '已到位': 2}
    df_output['_sort'] = df_output['到位状态'].map(status_order)
    df_output = df_output.sort_values(['_sort', '承诺收货日期']).drop(columns=['_sort'])

    # 10. 写入 Excel
    output_file = os.path.join(app.config['OUTPUT_FOLDER'], f'采购到位率分析结果_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: 到位率明细
        df_output.to_excel(writer, sheet_name='到位率明细', index=False)

        # Sheet 2: 汇总统计
        summary_data = []
        for status in ['未到位', '部分未到位', '已到位']:
            subset = df_output[df_output['到位状态'] == status]
            summary_data.append({
                '到位状态': status,
                '行数': len(subset),
                '采购数量合计': int(subset['采购数量'].sum()),
                '累计实收数量合计': int(subset['累计实收数量'].sum()),
                '未到位数量合计': int(subset['未到位数量'].sum()),
                '综合到位率': round(
                    subset['累计实收数量'].sum() / subset['采购数量'].sum()
                    if subset['采购数量'].sum() > 0 else 0, 4
                ),
            })
        summary_data.append({
            '到位状态': '合计',
            '行数': len(df_output),
            '采购数量合计': int(df_output['采购数量'].sum()),
            '累计实收数量合计': int(df_output['累计实收数量'].sum()),
            '未到位数量合计': int(df_output['未到位数量'].sum()),
            '综合到位率': round(
                df_output['累计实收数量'].sum() / df_output['采购数量'].sum()
                if df_output['采购数量'].sum() > 0 else 0, 4
            ),
        })
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='汇总统计', index=False)

        # Sheet 3: 匹配失败数据
        if len(df_unmatched) > 0:
            unmatched_cols = {
                'U9采购单号': 'U9采购单号',
                'U9采购单行号': 'U9采购单行号',
                '承诺收货日期': '承诺收货日期',
                '承诺收货数量': '承诺收货数量',
                '业务员姓名': '业务员',
                '物料规格': '物料规格',
                '物料型号': '物料型号',
                '物料代码': '物料代码',
            }
            existing_unmatched = {k: v for k, v in unmatched_cols.items() if k in df_unmatched.columns}
            df_unmatched_out = df_unmatched[list(existing_unmatched.keys())].rename(columns=existing_unmatched)
            df_unmatched_out.to_excel(writer, sheet_name='匹配失败数据', index=False)

        # Sheet 4: 到期履约情况
        today = pd.Timestamp(datetime.now().date())
        df_overdue = df_output[
            (df_output['承诺收货日期'].notna()) &
            (df_output['承诺收货日期'] <= today)
        ].copy()

        if len(df_overdue) > 0:
            overdue_sort_map = {'未到位': 0, '部分未到位': 1, '已到位': 2}
            df_overdue['_sort'] = df_overdue['到位状态'].map(overdue_sort_map)
            df_overdue = df_overdue.sort_values(['_sort', '承诺收货日期']).drop(columns=['_sort'])
            df_overdue.to_excel(writer, sheet_name='到期履约情况', index=False)

            overdue_summary = []
            for status in ['未到位', '部分未到位', '已到位']:
                subset = df_overdue[df_overdue['到位状态'] == status]
                overdue_summary.append({
                    '到位状态': status,
                    '行数': len(subset),
                    '采购数量合计': int(subset['采购数量'].sum()),
                    '累计实收数量合计': int(subset['累计实收数量'].sum()),
                    '未到位数量合计': int(subset['未到位数量'].sum()),
                    '综合到位率': round(
                        subset['累计实收数量'].sum() / subset['采购数量'].sum()
                        if subset['采购数量'].sum() > 0 else 0, 4
                    ),
                })
            overdue_summary.append({
                '到位状态': '合计',
                '行数': len(df_overdue),
                '采购数量合计': int(df_overdue['采购数量'].sum()),
                '累计实收数量合计': int(df_overdue['累计实收数量'].sum()),
                '未到位数量合计': int(df_overdue['未到位数量'].sum()),
                '综合到位率': round(
                    df_overdue['累计实收数量'].sum() / df_overdue['采购数量'].sum()
                    if df_overdue['采购数量'].sum() > 0 else 0, 4
                ),
            })
            df_overdue_summary = pd.DataFrame(overdue_summary)
            df_overdue_summary.to_excel(writer, sheet_name='到期履约汇总', index=False)

    # 11. 美化格式
    format_excel(output_file)

    # 12. 构建返回数据
    # 格式化汇总表数据（确保所有数值转为 Python 原生类型）
    def to_native(v):
        """numpy int/float → Python native"""
        if hasattr(v, 'item'):
            return v.item()
        return v

    def fmt_row(row_dict):
        return {
            '到位状态': str(row_dict['到位状态']),
            '行数': to_native(row_dict['行数']),
            '采购数量合计': to_native(row_dict.get('采购数量合计', 0)),
            '累计实收数量合计': to_native(row_dict.get('累计实收数量合计', 0)),
            '未到位数量合计': to_native(row_dict.get('未到位数量合计', 0)),
            '综合到位率': float(row_dict.get('综合到位率', 0)),
        }

    summary = {
        'total': to_native(len(df_output)),
        'matched': to_native(matched_count),
        'unmatched': to_native(unmatched_count),
        'not_arrived': to_native(summary_data[0]['行数']),
        'partial': to_native(summary_data[1]['行数']),
        'arrived': to_native(summary_data[2]['行数']),
        'total_qty': to_native(summary_data[3]['采购数量合计']),
        'total_received': to_native(summary_data[3]['累计实收数量合计']),
        'total_unreceived': to_native(summary_data[3]['未到位数量合计']),
        'overall_rate': f"{float(summary_data[3]['综合到位率']):.2%}",
        'detail_rows': [fmt_row(r) for r in summary_data],
    }

    overdue_info = {
        'count': to_native(len(df_overdue)),
        'detail_rows': [fmt_row(r) for r in overdue_summary] if len(df_overdue) > 0 else [],
    }

    return output_file, summary, overdue_info


def format_excel(filepath):
    """美化 Excel 格式"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(filepath)

    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    status_fills = {
        '未到位': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        '部分未到位': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        '已到位': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        # 列宽
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_text = ws.cell(row=1, column=col_idx).value
            ws.column_dimensions[col_letter].width = 16 if (header_text and len(str(header_text)) > 6) else 14

        # 表头
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 找"到位状态"列
        status_col = None
        rate_col = None
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val == '到位状态':
                status_col = col_idx
            if val in ('到位率', '综合到位率'):
                rate_col = col_idx

        # 数据行
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

            if status_col:
                status_val = str(ws.cell(row=row_idx, column=status_col).value or '')
                if status_val in status_fills:
                    ws.cell(row=row_idx, column=status_col).fill = status_fills[status_val]

            if rate_col:
                ws.cell(row=row_idx, column=rate_col).number_format = '0.00%'

        ws.freeze_panes = 'A2'
        if ws.max_row > 1:
            ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

    wb.save(filepath)


# ============================================================
# Web 路由
# ============================================================

@app.route('/')
def index():
    """主页面"""
    return render_template('index.html')


@app.route('/analyze', methods=['POST'])
def do_analyze():
    """执行分析"""
    try:
        # 获取上传文件
        pr_file = request.files.get('pr_file')
        po_file = request.files.get('po_file')
        commit_file = request.files.get('commit_file')

        if not all([pr_file, po_file, commit_file]):
            return jsonify({'success': False, 'error': '请上传全部3个文件'})

        # 保存上传文件
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        pr_path = os.path.join(app.config['UPLOAD_FOLDER'], f'PR_{timestamp}.xlsx')
        po_path = os.path.join(app.config['UPLOAD_FOLDER'], f'PO_{timestamp}.xls')
        commit_path = os.path.join(app.config['UPLOAD_FOLDER'], f'COMMIT_{timestamp}.xlsx')

        pr_file.save(pr_path)
        po_file.save(po_path)
        commit_file.save(commit_path)

        # 执行分析
        output_path, summary, overdue_info = analyze(pr_path, po_path, commit_path)

        # 返回结果
        filename = os.path.basename(output_path)
        return jsonify({
            'success': True,
            'summary': summary,
            'overdue': overdue_info,
            'download_url': f'/download/{filename}',
            'filename': filename,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'分析出错：{str(e)}'})


@app.route('/download/<filename>')
def download(filename):
    """下载结果文件"""
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True, download_name=filename)
    return '文件不存在', 404


@app.route('/cleanup', methods=['POST'])
def cleanup():
    """清理旧文件"""
    try:
        for folder in [app.config['UPLOAD_FOLDER'], app.config['OUTPUT_FOLDER']]:
            for f in os.listdir(folder):
                filepath = os.path.join(folder, f)
                if os.path.isfile(filepath):
                    os.remove(filepath)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


if __name__ == '__main__':
    print("=" * 60)
    print("  采购到位率查询系统 v1.0")
    print("  在浏览器中打开: http://127.0.0.1:5000")
    print("  按 Ctrl+C 停止服务")
    print("=" * 60)
    app.run(host='0.0.0.0', port=5000, debug=False)
